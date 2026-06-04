"""대선(대통령선거) 개표결과 Excel(data.nec.go.kr, 무인증) -> SQLite.

16~21대(2002~2025) 투표구별 Excel. 시도/시군구(일반구는 시로 합산)/투표구 단위 후보별 득표.
대선은 후보가 전국 동일 -> pres_cand(대수별 1세트). 시군구는 match_sigungu로 시 단위 합산.
포맷 상이: 헤더행/라벨행 위치, _x000D_ 잡음, 콤마 유무, 시도 short/full 혼재 모두 대응.
실행: python backend/data_pipeline/ingest_president.py [대수]   (기본 전부)
"""
import json
import re
import sqlite3
import pathlib
import sys

sys.path.insert(0, str(pathlib.Path(__file__).resolve().parent))
from ingest_council import match_sigungu, build_region_list, SIDO_FULL2SHORT  # noqa: E402

ROOT = pathlib.Path(__file__).resolve().parent.parent.parent
DB = ROOT / "backend" / "db" / "election.sqlite"
RAW = ROOT / "data" / "raw"
DAESU = [16, 17, 18, 19, 20, 21]
YEAR = {16: 2002, 17: 2007, 18: 2012, 19: 2017, 20: 2022, 21: 2025}
SIDO_CODE = {
    "서울": "11", "부산": "26", "대구": "27", "인천": "28", "광주": "29", "대전": "30",
    "울산": "31", "세종": "36", "경기": "41", "강원": "42", "충북": "43", "충남": "44",
    "전북": "45", "전남": "46", "경북": "47", "경남": "48", "제주": "50",
}
SPECIAL = {"거소투표", "거소·선상투표", "관외사전투표", "재외투표", "부재자투표",
           "국외부재자투표", "관내사전투표", "잠정투표"}
STOP = ("계", "무효", "기권", "후보자별")


def num(v):
    if v in (None, ""):
        return 0
    try:
        return int(str(v).replace(",", "").strip())
    except ValueError:
        return 0


def clean(s):
    if s is None:
        return ""
    return re.sub(r"_x000D_|\r", "", str(s)).strip()


def norm_sido(s):
    s = clean(s)
    if s in SIDO_CODE:
        return s
    return SIDO_FULL2SHORT.get(s)


def parse(daesu):
    import openpyxl
    wb = openpyxl.load_workbook(RAW / f"pres_{daesu}.xlsx", read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))

    # 헤더행 H = col0 가 시도/시도명
    H = next(i for i, r in enumerate(rows) if clean(r[0]) in ("시도", "시도명"))
    label_i = H + 1 if clean(rows[H][6]).startswith("후보자별") else H
    lbl = rows[label_i]
    cands = []
    for j in range(6, len(lbl)):
        t = clean(lbl[j])
        if not t or any(k in t for k in STOP):
            break
        parts = [p for p in t.split("\n") if p.strip()]
        party = parts[0].strip()
        name = parts[1].strip() if len(parts) > 1 else ""
        cands.append((party, name))
    ncand = len(cands)
    ccols = range(6, 6 + ncand)

    sido_tot = {}   # sido_short -> [votes]
    sgg_tot = {}    # code -> [votes] (시 단위 합산)
    precincts = []  # (code, dong, unit, tusu, [votes])
    cur_sido = cur_gu = cur_dong = None

    for r in rows[label_i + 1:]:
        s0, gu, dong, tu = clean(r[0]), clean(r[1]), clean(r[2]), clean(r[3])
        if s0:
            cur_sido = s0
        if gu:
            cur_gu = gu
            cur_dong = None
        sido = norm_sido(cur_sido)
        if cur_sido == "전국" or not gu and not dong and not tu:
            continue
        votes = [num(r[c]) for c in ccols]

        if gu.startswith("합계"):           # 시도 합계
            if sido:
                sido_tot[sido] = votes
            continue
        if not sido:
            continue
        code = None
        if cur_gu and not cur_gu.startswith("합계"):
            code, _ = match_sigungu(rlist, sido, cur_gu)

        if dong in ("합계", "계") and not tu:   # 시군구 합계
            if code:
                acc = sgg_tot.setdefault(code, [0] * ncand)
                for i, v in enumerate(votes):
                    acc[i] += v
            continue
        # 투표구 단위 (특수투표 + 일반 투표구). 소계는 동 carry용.
        if dong and tu == "소계":
            cur_dong = dong
            continue
        if dong in SPECIAL or tu or (dong and dong not in ("합계", "계")):
            d = dong if dong and dong not in SPECIAL else (cur_dong or "")
            if dong in SPECIAL:
                unit = dong
            else:
                unit = tu or dong
            if code and unit:
                precincts.append((code, d, unit, num(r[5]), votes))
    return cands, sido_tot, sgg_tot, precincts


def main():
    targets = [int(sys.argv[1])] if len(sys.argv) > 1 else DAESU
    con = sqlite3.connect(DB)
    cur = con.cursor()
    cur.executescript("""
    CREATE TABLE IF NOT EXISTS pres_cand(daesu INT, idx INT, party TEXT, name TEXT);
    CREATE TABLE IF NOT EXISTS pres_region(daesu INT, level TEXT, region_code TEXT, idx INT, votes INT, rate REAL);
    CREATE TABLE IF NOT EXISTS pres_precinct(daesu INT, region_code TEXT, dong TEXT, unit TEXT, tusu INT, votes_json TEXT);
    CREATE INDEX IF NOT EXISTS idx_presreg ON pres_region(daesu, level, region_code);
    CREATE INDEX IF NOT EXISTS idx_prespre ON pres_precinct(daesu, region_code);
    """)
    global rlist
    rlist = build_region_list(con)

    for daesu in targets:
        cur.execute("DELETE FROM pres_cand WHERE daesu=?", (daesu,))
        cur.execute("DELETE FROM pres_region WHERE daesu=?", (daesu,))
        cur.execute("DELETE FROM pres_precinct WHERE daesu=?", (daesu,))
        cands, sido_tot, sgg_tot, precincts = parse(daesu)
        for i, (party, name) in enumerate(cands):
            cur.execute("INSERT INTO pres_cand VALUES(?,?,?,?)", (daesu, i, party, name))

        def ins_region(level, code, votes):
            valid = sum(votes) or 1
            for i, v in enumerate(votes):
                cur.execute("INSERT INTO pres_region VALUES(?,?,?,?,?,?)",
                            (daesu, level, code, i, v, round(v / valid * 100, 2)))

        for sido, votes in sido_tot.items():
            ins_region("시도", SIDO_CODE[sido], votes)
        for code, votes in sgg_tot.items():
            ins_region("구시군", code, votes)
        for code, dong, unit, tusu, votes in precincts:
            cur.execute("INSERT INTO pres_precinct VALUES(?,?,?,?,?,?)",
                        (daesu, code, dong, unit, tusu, json.dumps(votes)))
        con.commit()
        win = max(range(len(cands)), key=lambda i: sido_tot.get("서울", [0] * len(cands))[i]) if cands else 0
        print(f"{daesu}대({YEAR[daesu]}): 후보 {len(cands)} | 시도 {len(sido_tot)} | 시군구 {len(sgg_tot)} | 투표구 {len(precincts)}")
        print(f"   후보: {[f'{p}/{n}' for p, n in cands]}")
    con.close()


if __name__ == "__main__":
    main()
