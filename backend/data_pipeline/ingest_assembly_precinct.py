"""총선 투표구별 개표결과 Excel(data.nec.go.kr) -> SQLite.

선거구별 후보 전원(낙선 포함) + 투표구별 득표.
Excel 선거구명이 geo 약칭과 달라 '당선자 이름'으로 geo key(SIDO_SGG) 매칭
(assembly_{대수}_names.json 사용).
실행: python backend/data_pipeline/ingest_assembly_precinct.py <대수>
"""
import json
import sqlite3
import pathlib
import sys

ROOT = pathlib.Path(__file__).resolve().parent.parent.parent
DB = ROOT / "backend" / "db" / "election.sqlite"
RAW = ROOT / "data" / "raw"

FILE = {22: "assembly_22_precinct.xlsx", 21: "assembly_21_precinct.xlsx"}
SIDO_FULL2SHORT = {
    "서울특별시": "서울", "부산광역시": "부산", "대구광역시": "대구", "인천광역시": "인천",
    "광주광역시": "광주", "대전광역시": "대전", "울산광역시": "울산", "세종특별자치시": "세종",
    "경기도": "경기", "강원도": "강원", "강원특별자치도": "강원", "충청북도": "충북",
    "충청남도": "충남", "전라북도": "전북", "전북특별자치도": "전북", "전라남도": "전남",
    "경상북도": "경북", "경상남도": "경남", "제주특별자치도": "제주",
}


def num(v):
    if v is None or v == "":
        return None
    try:
        return int(str(v).replace(",", "").strip())
    except ValueError:
        return None


def parse(daesu):
    import openpyxl
    wb = openpyxl.load_workbook(RAW / FILE[daesu], read_only=True)
    ws = wb["지역구"]
    width = ws.max_column
    cand_cols = list(range(6, width - 3))  # 후보 컬럼들 (계/무효/기권 앞)

    blocks = {}  # (sido_short, sgg_name) -> {parties, names, total[], precincts:[...]}
    cur = None
    expect = None  # 'party' or 'cand'
    parties = names = None

    for r in ws.iter_rows(min_row=2, values_only=True):
        sido, sgg = r[0], r[1]
        eup, tu, seonin = r[2], r[3], r[4]
        is_header = (seonin in (None, "")) and (r[6] not in (None, "")) and (eup in (None, ""))
        if is_header:
            if expect != "cand":
                parties = [r[c] for c in cand_cols]
                expect = "cand"
            else:
                names = [r[c] for c in cand_cols]
                key = (SIDO_FULL2SHORT.get(sido, sido), sgg)
                cand = [(parties[i], names[i]) for i in range(len(cand_cols)) if names[i] not in (None, "")]
                blocks[key] = {"cand": cand, "ncand": len(cand), "total": None, "precincts": []}
                cur = key
                expect = None
            continue
        if cur is None or num(seonin) is None:
            continue
        votes = [num(r[c]) or 0 for c in cand_cols[: blocks[cur]["ncand"]]]
        tusu = num(r[5])
        if eup == "합계":
            blocks[cur]["total"] = {"votes": votes, "tusu": tusu, "seonin": num(seonin)}
        else:
            # 투표구 단위 행(거소/관외/국외/관내사전/일반 투표구). 동명 carry.
            dong = eup if eup not in (None, "") else blocks[cur].get("_dong", "")
            if eup not in (None, "") and tu == "소계":
                blocks[cur]["_dong"] = eup
                continue  # 소계는 동 carry용, 표엔 투표구만
            unit = tu if tu not in (None, "") else eup
            blocks[cur]["precincts"].append({"dong": dong, "unit": unit, "tusu": tusu, "votes": votes})
    return blocks


def main():
    daesu = int(sys.argv[1]) if len(sys.argv) > 1 else 22
    names = json.loads((RAW / f"assembly_{daesu}_names.json").read_text(encoding="utf-8"))
    name2key = {}
    for sidosgg, nm in names.items():
        sido = sidosgg.split(" ", 1)[0]
        name2key[(sido, nm)] = sidosgg

    blocks = parse(daesu)
    con = sqlite3.connect(DB)
    cur = con.cursor()
    cur.executescript("""
    CREATE TABLE IF NOT EXISTS gukhoe_cand(
      daesu INT, key TEXT, idx INT, party TEXT, name TEXT, votes INT, rate REAL, elected INT);
    CREATE TABLE IF NOT EXISTS gukhoe_precinct(
      daesu INT, key TEXT, dong TEXT, unit TEXT, tusu INT, votes_json TEXT);
    CREATE INDEX IF NOT EXISTS idx_gcand ON gukhoe_cand(daesu, key);
    CREATE INDEX IF NOT EXISTS idx_gprec ON gukhoe_precinct(daesu, key);
    """)
    cur.execute("DELETE FROM gukhoe_cand WHERE daesu=?", (daesu,))
    cur.execute("DELETE FROM gukhoe_precinct WHERE daesu=?", (daesu,))

    matched = 0
    misses = []
    for (sido, sgg), b in blocks.items():
        if not b["total"]:
            continue
        tv = b["total"]["votes"]
        win_i = max(range(len(tv)), key=lambda i: tv[i]) if tv else 0
        win_name = b["cand"][win_i][1]
        key = name2key.get((sido, win_name))
        if not key:
            misses.append((sido, sgg, win_name))
            continue
        matched += 1
        valid = sum(tv) or 1
        for i, (party, name) in enumerate(b["cand"]):
            cur.execute("INSERT INTO gukhoe_cand VALUES(?,?,?,?,?,?,?,?)",
                        (daesu, key, i, party, name, tv[i], round(tv[i] / valid * 100, 2), 1 if i == win_i else 0))
        for p in b["precincts"]:
            cur.execute("INSERT INTO gukhoe_precinct VALUES(?,?,?,?,?,?)",
                        (daesu, key, p["dong"], p["unit"], p["tusu"], json.dumps(p["votes"])))
    con.commit()
    print(f"{daesu}대: 선거구 {len(blocks)} / 매칭 {matched}")
    if misses:
        print("  미매칭:", misses[:10], f"(총 {len(misses)})")
    print(f"  gukhoe_cand: {cur.execute('SELECT COUNT(*) FROM gukhoe_cand WHERE daesu=?', (daesu,)).fetchone()[0]}")
    print(f"  gukhoe_precinct: {cur.execute('SELECT COUNT(*) FROM gukhoe_precinct WHERE daesu=?', (daesu,)).fetchone()[0]}")
    con.close()


if __name__ == "__main__":
    main()
